import openpyxl
from tkinter import *


class Exportar_Info:

    def __init__(self, janela_incio, db_consulta):
        self.janela_inicio = janela_incio
        self.db_consulta = db_consulta

        self.mensagem = Label(self.janela_inicio, text='', fg='red', font=('Arial', 15))
        self.mensagem.place(x=700, y=30)

    def exportar_excel(self):
        query_veiculos = 'SELECT * FROM Veiculos'
        registos_veiculos = self.db_consulta(query_veiculos)

        query_clientes = 'SELECT * FROM Clientes'
        registos_clientes = self.db_consulta(query_clientes)

        query_utilizadores = 'SELECT * FROM Utilizador'
        registos_utilizadores = self.db_consulta(query_utilizadores)

        query_reservas = 'SELECT * FROM Reservas'
        registos_reservas = self.db_consulta(query_reservas)

        wb = openpyxl.load_workbook('exportações/exportacoes_luxury_excel.xlsx')

        if 'Reservas' in wb.sheetnames:
            folha_reservas = wb['Reservas']
        else:
            folha_reservas = wb.create_sheet('Reservas')

        folha_reservas.delete_rows(2, folha_reservas.max_row)  # Apagar as linhas antigas antes de adicionar as novas

        for linha in registos_reservas:
            folha_reservas['A1'] = 'ID'
            folha_reservas['B1'] = 'Nome'
            folha_reservas['C1'] = 'Marca'
            folha_reservas['D1'] = 'Modelo'
            folha_reservas['E1'] = 'Data Inicio'
            folha_reservas['F1'] = 'Data Fim'
            folha_reservas['G1'] = 'Preço Dia'
            folha_reservas['H1'] = 'Preço Total'
            folha_reservas.append(linha)

        if 'Clientes' in wb.sheetnames:
            folha_clientes = wb['Clientes']
        else:
            folha_clientes = wb.create_sheet('Clientes')

        folha_clientes.delete_rows(2, folha_clientes.max_row)

        for linha in registos_clientes:
            folha_clientes['A1'] = 'ID'
            folha_clientes['B1'] = 'Nome'
            folha_clientes['C1'] = 'Password'
            folha_clientes['D1'] = 'Email'
            folha_clientes['E1'] = 'Telemóvel'
            folha_clientes['F1'] = 'Cartão Bancário'
            folha_clientes['G1'] = 'Data Inscrição'
            folha_clientes.append(linha)

        if 'Veiculos' in wb.sheetnames:
            folha_veiculos = wb['Veiculos']
        else:
            folha_veiculos = wb.create_sheet('Veiculos')

        folha_veiculos.delete_rows(2, folha_veiculos.max_row)

        for linha in registos_veiculos:
            folha_veiculos['A1'] = 'ID'
            folha_veiculos['B1'] = 'Marca'
            folha_veiculos['C1'] = 'Modelo'
            folha_veiculos['D1'] = 'Categoria'
            folha_veiculos['E1'] = 'Tipo'
            folha_veiculos['F1'] = 'Quantidade Pessoas'
            folha_veiculos['G1'] = 'Imagem'
            folha_veiculos['H1'] = 'Preço'
            folha_veiculos['I1'] = 'Ultima Revisão'
            folha_veiculos['J1'] = 'Ultima Legalização'
            folha_veiculos['K1'] = 'Manutenção'
            folha_veiculos['L1'] = 'Disponivel'
            folha_veiculos.append(linha)

        if 'Utilizadores' in wb.sheetnames:
            folha_utilizadores = wb['Utilizadores']
        else:
            folha_utilizadores = wb.create_sheet('Utilizadores')

        folha_utilizadores.delete_rows(2, folha_utilizadores.max_row)

        for linha in registos_utilizadores:
            folha_utilizadores['A1'] = 'ID'
            folha_utilizadores['B1'] = 'Nome'
            folha_utilizadores['C1'] = 'Username'
            folha_utilizadores['D1'] = 'Password'
            folha_utilizadores['E1'] = 'Email'
            folha_utilizadores['F1'] = 'Telemóvel'
            folha_utilizadores['G1'] = 'Admin'
            folha_utilizadores.append(linha)

        wb.save('exportações/exportacoes_luxury_excel.xlsx')

        self.mensagem['text'] = 'Informações exportadas para EXCEL com sucesso!'
        self.janela_inicio.after(1500, lambda: self.mensagem.config(text=''))